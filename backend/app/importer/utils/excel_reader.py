from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import load_workbook


def read_excel_rows(file_bytes: bytes, sheet_name: str = "Grid 1") -> list[dict[str, Any]]:
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Worksheet '{sheet_name}' not found.")
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    result: list[dict[str, Any]] = []
    for idx, row in enumerate(rows[1:], start=2):
        row_dict: dict[str, Any] = {"__rownum__": idx}
        for col_idx, header in enumerate(headers):
            if not header:
                continue
            row_dict[header] = row[col_idx] if col_idx < len(row) else None
        result.append(row_dict)
    return result
